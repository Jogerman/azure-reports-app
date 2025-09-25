# backend/apps/reports/utils/export_extensions.py
"""
Extensiones de exportación para reportes especializados
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
import io
import json
from datetime import datetime
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)

class ExportManager:
    """Gestor de exportaciones en múltiples formatos"""
    
    @staticmethod
    def export_to_excel(analysis_data: Dict[str, Any], report_type: str, client_name: str = "Cliente") -> bytes:
        """Exportar análisis a Excel con formato profesional"""
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            
            # Remover hoja por defecto
            wb.remove(wb.active)
            
            # Crear hojas según el tipo de reporte
            if report_type == 'security':
                ExportManager._create_security_excel_sheets(wb, analysis_data, client_name)
            elif report_type == 'performance':
                ExportManager._create_performance_excel_sheets(wb, analysis_data, client_name)
            elif report_type == 'cost':
                ExportManager._create_cost_excel_sheets(wb, analysis_data, client_name)
            else:
                ExportManager._create_comprehensive_excel_sheets(wb, analysis_data, client_name)
            
            # Guardar en bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            raise
    
    @staticmethod
    def _create_security_excel_sheets(wb: openpyxl.Workbook, analysis_data: Dict, client_name: str):
        """Crear hojas Excel para reporte de seguridad"""
        # Hoja 1: Resumen Ejecutivo
        summary_sheet = wb.create_sheet("Resumen Ejecutivo")
        
        # Headers y título
        summary_sheet['A1'] = f"Análisis de Seguridad - {client_name}"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        summary_sheet['A1'].font = Font(color='FFFFFF', size=16, bold=True)
        
        # Métricas principales
        dashboard_metrics = analysis_data.get('dashboard_metrics', {})
        basic_metrics = analysis_data.get('basic_metrics', {})
        
        row = 3
        metrics = [
            ('Security Score', dashboard_metrics.get('security_score', 0)),
            ('Total Acciones de Seguridad', dashboard_metrics.get('total_actions', 0)),
            ('Issues Críticos', dashboard_metrics.get('critical_issues', 0)),
            ('Horas de Trabajo Estimadas', dashboard_metrics.get('working_hours', 0)),
            ('Nivel de Riesgo', dashboard_metrics.get('risk_level', 'Unknown'))
        ]
        
        for metric_name, metric_value in metrics:
            summary_sheet[f'A{row}'] = metric_name
            summary_sheet[f'B{row}'] = metric_value
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Hoja 2: Compliance Gaps
        compliance_sheet = wb.create_sheet("Compliance Gaps")
        compliance_sheet['A1'] = "Análisis de Gaps de Cumplimiento"
        compliance_sheet['A1'].font = Font(size=14, bold=True)
        
        compliance_gaps = analysis_data.get('compliance_gaps', {})
        row = 3
        compliance_sheet['A2'] = "Tipo de Gap"
        compliance_sheet['B2'] = "Cantidad"
        
        for gap_type, count in compliance_gaps.items():
            compliance_sheet[f'A{row}'] = gap_type.replace('_', ' ').title()
            compliance_sheet[f'B{row}'] = count
            row += 1
        
        # Hoja 3: Recomendaciones Prioritarias
        recommendations_sheet = wb.create_sheet("Recomendaciones")
        recommendations_sheet['A1'] = "Recomendaciones Prioritarias"
        recommendations_sheet['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Tipo de Recurso', 'Recomendación', 'Impacto de Negocio']
        for col, header in enumerate(headers, 1):
            cell = recommendations_sheet.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Datos
        priority_recs = analysis_data.get('priority_recommendations', [])
        for row, rec in enumerate(priority_recs, 3):
            recommendations_sheet[f'A{row}'] = rec.get('resource_type', '')
            recommendations_sheet[f'B{row}'] = rec.get('recommendation', '')
            recommendations_sheet[f'C{row}'] = rec.get('business_impact', '')
        
        # Autofit columns
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _create_performance_excel_sheets(wb: openpyxl.Workbook, analysis_data: Dict, client_name: str):
        """Crear hojas Excel para reporte de rendimiento"""
        # Implementación similar a security pero con métricas de performance
        summary_sheet = wb.create_sheet("Resumen Performance")
        summary_sheet['A1'] = f"Análisis de Rendimiento - {client_name}"
        summary_sheet['A1'].font = Font(size=16, bold=True, color='FF6600')
        
        dashboard_metrics = analysis_data.get('dashboard_metrics', {})
        
        metrics = [
            ('Performance Score', dashboard_metrics.get('performance_score', 100)),
            ('Total Optimizaciones', dashboard_metrics.get('total_actions', 0)),
            ('Optimizaciones Críticas', dashboard_metrics.get('critical_optimizations', 0)),
            ('Potencial de Mejora (%)', dashboard_metrics.get('optimization_potential', 0)),
            ('Calificación de Eficiencia', dashboard_metrics.get('efficiency_rating', 'Good'))
        ]
        
        for row, (metric_name, metric_value) in enumerate(metrics, 3):
            summary_sheet[f'A{row}'] = metric_name
            summary_sheet[f'B{row}'] = metric_value
            summary_sheet[f'A{row}'].font = Font(bold=True)
    
    @staticmethod
    def _create_cost_excel_sheets(wb: openpyxl.Workbook, analysis_data: Dict, client_name: str):
        """Crear hojas Excel para reporte de costos"""
        # Hoja de resumen de costos
        summary_sheet = wb.create_sheet("Resumen Costos")
        summary_sheet['A1'] = f"Análisis de Costos - {client_name}"
        summary_sheet['A1'].font = Font(size=16, bold=True, color='009900')
        
        dashboard_metrics = analysis_data.get('dashboard_metrics', {})
        
        # Métricas principales
        metrics = [
            ('Ahorros Mensuales Estimados', f"${dashboard_metrics.get('monthly_savings', 0):,.0f}"),
            ('Ahorros Anuales Estimados', f"${dashboard_metrics.get('annual_savings', 0):,.0f}"),
            ('ROI Mensual (%)', f"{dashboard_metrics.get('roi_percentage', 0):.1f}%"),
            ('Período de Recuperación (meses)', dashboard_metrics.get('payback_months', 0)),
            ('Score de Optimización', f"{dashboard_metrics.get('optimization_score', 100)}%")
        ]
        
        for row, (metric_name, metric_value) in enumerate(metrics, 3):
            summary_sheet[f'A{row}'] = metric_name
            summary_sheet[f'B{row}'] = metric_value
            summary_sheet[f'A{row}'].font = Font(bold=True)
        
        # Hoja de desglose de ahorros
        savings_sheet = wb.create_sheet("Desglose Ahorros")
        savings_analysis = analysis_data.get('savings_analysis', {})
        
        savings_breakdown = [
            ('Ahorros Inmediatos', savings_analysis.get('immediate_savings', 0)),
            ('Ahorros Corto Plazo', savings_analysis.get('short_term_savings', 0)),
            ('Ahorros Largo Plazo', savings_analysis.get('long_term_savings', 0))
        ]
        
        for row, (category, amount) in enumerate(savings_breakdown, 2):
            savings_sheet[f'A{row}'] = category
            savings_sheet[f'B{row}'] = f"${amount:,.0f}"
    
    @staticmethod
    def _create_comprehensive_excel_sheets(wb: openpyxl.Workbook, analysis_data: Dict, client_name: str):
        """Crear hojas Excel para reporte completo"""
        # Implementar para reporte comprehensivo
        summary_sheet = wb.create_sheet("Resumen General")
        summary_sheet['A1'] = f"Análisis Completo Azure - {client_name}"
        summary_sheet['A1'].font = Font(size=16, bold=True)
    
    @staticmethod
    def export_to_json(analysis_data: Dict[str, Any], report_info: Dict) -> str:
        """Exportar análisis a JSON estructurado"""
        try:
            export_data = {
                'report_metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'report_type': report_info.get('type', 'unknown'),
                    'report_title': report_info.get('title', ''),
                    'client_name': report_info.get('client_name', 'Cliente'),
                    'version': '1.0'
                },
                'analysis_data': analysis_data,
                'export_format': 'json_v1'
            }
            
            return json.dumps(export_data, indent=2, ensure_ascii=False)
            
        except Exception as e:
            logger.error(f"Error exportando a JSON: {e}")
            raise
    
    @staticmethod
    def export_to_csv_summary(analysis_data: Dict[str, Any], report_type: str) -> str:
        """Exportar resumen como CSV"""
        try:
            dashboard_metrics = analysis_data.get('dashboard_metrics', {})
            
            if report_type == 'security':
                data = [
                    ['Metric', 'Value'],
                    ['Security Score', dashboard_metrics.get('security_score', 0)],
                    ['Total Actions', dashboard_metrics.get('total_actions', 0)],
                    ['Critical Issues', dashboard_metrics.get('critical_issues', 0)],
                    ['Working Hours', dashboard_metrics.get('working_hours', 0)],
                    ['Risk Level', dashboard_metrics.get('risk_level', 'Unknown')]
                ]
            elif report_type == 'performance':
                data = [
                    ['Metric', 'Value'],
                    ['Performance Score', dashboard_metrics.get('performance_score', 100)],
                    ['Total Actions', dashboard_metrics.get('total_actions', 0)],
                    ['Critical Optimizations', dashboard_metrics.get('critical_optimizations', 0)],
                    ['Optimization Potential (%)', dashboard_metrics.get('optimization_potential', 0)],
                    ['Efficiency Rating', dashboard_metrics.get('efficiency_rating', 'Good')]
                ]
            elif report_type == 'cost':
                data = [
                    ['Metric', 'Value'],
                    ['Monthly Savings', f"${dashboard_metrics.get('monthly_savings', 0):,.0f}"],
                    ['Annual Savings', f"${dashboard_metrics.get('annual_savings', 0):,.0f}"],
                    ['ROI Percentage', f"{dashboard_metrics.get('roi_percentage', 0):.1f}%"],
                    ['Payback Months', dashboard_metrics.get('payback_months', 0)],
                    ['Optimization Score', f"{dashboard_metrics.get('optimization_score', 100)}%"]
                ]
            else:
                data = [['Metric', 'Value'], ['Total Actions', dashboard_metrics.get('total_actions', 0)]]
            
            # Convertir a CSV
            csv_output = []
            for row in data:
                csv_output.append(','.join([str(cell) for cell in row]))
            
            return '\n'.join(csv_output)
            
        except Exception as e:
            logger.error(f"Error exportando CSV: {e}")
            raise

